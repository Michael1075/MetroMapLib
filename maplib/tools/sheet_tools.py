import openpyxl

from maplib.tools.assertions import assert_type


def open_xlsx_file(file_name):
    return openpyxl.load_workbook(filename = file_name)


def get_sheet_val(sheet, min_row = None, max_row = None, min_column = None, max_column = None):
    assert_type(sheet, openpyxl.worksheet.worksheet.Worksheet)
    if min_row is None:
        min_row = 0
    if max_row is None:
        max_row = sheet.max_row
    if min_column is None:
        min_column = 0
    if max_column is None:
        max_column = sheet.max_column
    return tuple([
        tuple([
            sheet.cell(row_index + 1, column_index + 1).value
            for column_index in range(min_column, max_column)
        ])
        for row_index in range(min_row, max_row)
    ])


def get_row_val(sheet, row_index, min_column = None, max_column = None):
    result = get_sheet_val(sheet, row_index, row_index + 1, min_column, max_column)
    return result[0]


def get_column_val(sheet, column_index, min_row = None, max_row = None):
    result = get_sheet_val(sheet, min_row, max_row, column_index, column_index + 1)
    return tuple([val[0] for val in result])


def get_workbook_data_dict(file_name, ignore_none = False, child_sheet_max_column = None):
    result = dict()
    database = open_xlsx_file(file_name)
    main_sheet = database["main"]
    num_items = main_sheet.max_row
    for k in range(num_items):
        row_data = get_row_val(main_sheet, k)
        if ignore_none:
            row_data = [val for val in row_data if val is not None]
        sheet_name = row_data[0]
        basic_data = row_data[1:]
        detailed_data = get_sheet_val(database[sheet_name], max_column = child_sheet_max_column)
        result[sheet_name] = (basic_data, detailed_data)
    return result

