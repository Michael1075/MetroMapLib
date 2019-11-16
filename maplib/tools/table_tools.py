import openpyxl


def get_table_val(table, min_row = None, max_row = None, min_column = None, max_column = None):
    assert isinstance(table, openpyxl.worksheet.worksheet.Worksheet)
    if min_row is None:
        min_row = 0
    if max_row is None:
        max_row = table.max_row
    if min_column is None:
        min_column = 0
    if max_column is None:
        max_column = table.max_column
    return tuple([
        tuple([
            table.cell(row_index + 1, column_index + 1).value
            for column_index in range(min_column, max_column)
        ])
        for row_index in range(min_row, max_row)
    ])


def get_row_val(table, row_index, min_column = None, max_column = None):
    result = get_table_val(table, row_index, row_index + 1, min_column, max_column)
    return result[0]


def get_column_val(table, column_index, min_row = None, max_row = None):
    result = get_table_val(table, min_row, max_row, column_index, column_index + 1)
    return tuple([val[0] for val in result])

